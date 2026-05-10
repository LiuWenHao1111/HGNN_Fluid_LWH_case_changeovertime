import os
import pandas as pd
import csv, ast
from openpyxl import Workbook

class CsvToExcelConverter:
    """
    读取CSV文件并转换为Excel文件，确保数据和位置保持一致。
    """
    def __init__(self, path, file_name):
        self.path = path  # CSV文件路径
        self.file_name = file_name  # 文件名（无扩展名）
        self.excel_file_path = os.path.join(self.path, f"{self.file_name}.xlsx")

    def read_csv_files(self):
        """
        读取CSV文件并返回DataFrame。
        """
        base_data = os.path.join(self.path, self.file_name, 'based_data.csv')
        base_data_dict = self.read_csv_to_dict_base(base_data)
        process_data = os.path.join(self.path, self.file_name, 'process_data.csv')
        kind_machine_dict, kind_process_time_dict, machine_kind_dict, kind_task_number_dict\
            = self.read_csv_to_nested_dict_process(process_data)
        order_data = os.path.join(self.path, self.file_name, 'order_data.csv')
        order_data_dict = self.read_csv_to_dict_order(order_data)
        return (base_data_dict, kind_machine_dict, kind_process_time_dict,
                machine_kind_dict, kind_task_number_dict, order_data_dict)

    def read_csv_to_dict_base(self, file_path):
        with open(file_path, mode='r', newline='') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                base_info_dict = {
                    'machine_count': [int(row['machine_count'])],
                    'kind_count': [int(row['kind_count'])],
                    'order_count': [int(row['order_count'])],
                    'DDT': [float(row['DDT'])]
                }
                return base_info_dict

    def read_csv_to_dict_order(self, file_path):
        order_dict = {}
        with open(file_path, mode='r', newline='') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for row in csvreader:
                order = int(row['order'])
                order_dict[order] = {
                    'time_arrive': int(row['time_arrive']),
                    'time_delivery': int(row['time_delivery']),
                    'kind_number': ast.literal_eval(row['kind_number'])
                }
        return order_dict

    def read_csv_to_nested_dict_process(self, file_path):
        kind_machine_dict = {}
        kind_process_time_dict = {}
        machine_kind_dict = {}
        kind_task_number_dict = {}

        with open(file_path, mode='r', newline='') as csvfile:
            csvreader = csv.DictReader(csvfile)
            for line_number, row in enumerate(csvreader):
                machine_selectable = ast.literal_eval(row['machine_selectable'])
                process_time = ast.literal_eval(row['process_time'])
                kind = int(row['kind'])
                task = int(row['task'])
                kind_task_number_dict[kind] = task + 1 # 记录该类工件的工序数

                # 使用行号作为键
                kind_machine_dict[line_number] = machine_selectable
                kind_process_time_dict[line_number] = process_time

                # 更新机器的可加工类列表
                for machine in machine_selectable:
                    if machine not in machine_kind_dict:
                        machine_kind_dict[machine] = []
                    machine_kind_dict[machine].append(line_number)

        return kind_machine_dict, kind_process_time_dict, machine_kind_dict, kind_task_number_dict

    def write_to_excel(self):
        """
        将二维列表写入Excel文件。
        """
        # 读取数据
        base_data_dict, kind_machine_dict, kind_process_time_dict, machine_kind_dict, kind_task_number_dict, order_data_dict \
            = self.read_csv_files()

        kind_task_number_list = []
        for key in range(base_data_dict['kind_count'][0]):
            kind_task_number_list.append(kind_task_number_dict[key])

        wb = Workbook()

        # 填充算例基础数据: 第一列为：machine_count、kind_count、kind_task_number、order_count、time_arrive、kind_number
        # 对应列标后接对应的数据
        ws_base = wb.active
        ws_base.title = "算例基础数据"
        # 第一列按行写入
        for item1, item2 in zip(['machine_count', 'kind_count', 'kind_task_number', 'order_count', 'time_arrive', 'kind_number'],
                                [[base_data_dict['machine_count'][0]], [base_data_dict['kind_count'][0]], kind_task_number_list,
                                 [base_data_dict['order_count'][0]], [order_data_dict[0]['time_arrive']], order_data_dict[0]['kind_number']]
                                ):
            ws_base.append([item1] + list(item2))

        # 填充各类可选加工机器数据
        ws_process = wb.create_sheet(title='各类可选加工机器')
        for key in kind_machine_dict:
            ws_process.append(kind_machine_dict[key])

        # 填充各类在各机器加工时间数据
        ws_order = wb.create_sheet(title='各类在各机器加工时间')
        for key in kind_process_time_dict:
            ws_order.append(kind_process_time_dict[key])

        # 填充各机器可加工类数据
        ws_machine = wb.create_sheet(title='各机器可加工类')
        for key in range(base_data_dict['machine_count'][0]):
            ws_machine.append(machine_kind_dict[key])

        # 保存Excel文件
        wb.save(self.excel_file_path)
        print(f"Excel文件已生成：{self.excel_file_path}")


# 测试
if __name__ == '__main__':
    file_name_list = \
        ['M10R5N5', 'M10R5N10', 'M10R5N15', 'M10R10N5', 'M10R10N10', 'M10R10N15', 'M10R15N5', 'M10R15N10', 'M10R15N15',
         'M15R5N5', 'M15R5N10', 'M15R5N15', 'M15R10N5', 'M15R10N10', 'M15R10N15', 'M15R15N5', 'M15R15N10', 'M15R15N15',
         'M20R5N5', 'M20R5N10', 'M20R5N15', 'M20R10N5', 'M20R10N10', 'M20R10N15', 'M20R15N5', 'M20R15N10', 'M20R15N15']
    file_name_list = ['M10R5N5']
    path_save = 'D:/Python project/Deep_Reinforcement_Learning_FJSP/data/DSLFMAE'
    for file_name in file_name_list:
        converter = CsvToExcelConverter(path=path_save, file_name=file_name)
        converter.write_to_excel()
        print('转换完成文件：', file_name)
