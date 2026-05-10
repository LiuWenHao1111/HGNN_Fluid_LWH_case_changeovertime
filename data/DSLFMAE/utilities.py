import openpyxl

class DataRead():
    """
    excel数据读取: data
    """
    def __init__(self, book_name_xlsx, input_data, machine_class_data, time_machine_class_data, class_machine_data):
        # 初始化参数
        self.input_data = input_data
        self.machine_class_data = machine_class_data
        self.time_machine_class_data = time_machine_class_data
        self.class_machine_data = class_machine_data
        self.book_name_xlsx = book_name_xlsx
        # 读取输入数据[机器数，工件种类数，各类工件包含的工序列表，订单到达次数，各订单每次到达时间列表，各订单包含工件二维列表]
        [self.machine_number, self.number_job_kind, self.number_task, self.order_arrival_number, self.order_arrival_time, self.order_list] = \
            self.read_excel_input_data(self.input_data)
        self.machine_class = self.read_sheet(self.machine_class_data)      # 读取各类可选加工机器
        self.class_machine = self.read_sheet(self.class_machine_data)      # 读取各机器可选加工类
        self.time_machine_class = self.read_sheet(self.time_machine_class_data)    # 读取各类在各机器加工时间

    def read_excel_input_data(self, sheet_name):
        """
        读取对应输入表格数据
        :param path: excel文件名
        :param sheet_name: 表格名
        :return: 读取出的表格列表
        """
        wb = openpyxl.load_workbook(self.book_name_xlsx)  # 打开excel文件，获取工作薄对象
        ws = wb[sheet_name]  # 从表单中获取单元格的内容
        titles = ['机器数', '工件种类数', '每种工件工序数', '订单到达次数', '订单到达时间']
        dict_value = {}  # 返回值
        for row in range(len(titles)):    # [row: 机器数，工件种类数，每种工件工序数，订单到达次数，订单到达时间]
            values = []
            for col in range(1, ws.max_column):
                value = ws.cell(row=row + 1, column=col + 1).value
                if value is not None:
                    values.append(value)
            dict_value[titles[row]] = tuple(values)
        # 读取各到达订单包含的工件数量信息
        value_tuple = []
        for row in range(len(titles), ws.max_row):
            values = []
            for col in range(1, ws.max_column):
                value = ws.cell(row=row + 1, column=col + 1).value
                if value is not None:
                    values.append(value)
            value_tuple.append(tuple(values))
        dict_value['各订单包含的工件数'] = tuple(value_tuple)

        return [dict_value['机器数'][0], dict_value['工件种类数'][0], dict_value['每种工件工序数'], dict_value['订单到达次数'][0], dict_value['订单到达时间'], dict_value['各订单包含的工件数']]

    def read_sheet(self, sheet_name):
        """
        读取生成的各数据值
        :param sheet_name: sheet表
        :return: 数据值
        """
        wb = openpyxl.load_workbook(self.book_name_xlsx)  # 打开excel文件，获取工作薄对象
        ws = wb[sheet_name]  # 从表单中获取单元格的内容
        tuple_values = []  # 返回值
        for row in range(ws.max_row):
            values = []
            for col in range(ws.max_column):
                value = ws.cell(row=row + 1, column=col + 1).value
                if value is not None:
                    values.append(value)
            tuple_values.append(tuple(values))
        return tuple(tuple_values)

class DataComputer(DataRead):
    """
    数据计算索引类: indexed
    """
    def __init__(self, book_name_xlsx, input_data, machine_class_data, time_machine_class_data, class_machine_data):
        super().__init__(book_name_xlsx, input_data, machine_class_data, time_machine_class_data, class_machine_data)
        self.machine_list = tuple(range(self.machine_number))  # 机器列表
        self.number_class = sum(self.number_task)   # 类总数
        self.class_list = tuple(range(self.number_class))  # 类列表
        self.class_kind = self.class_kind_function(self)     # 各种工件对应的类字典，工件种类【包含类集合】
        self.kind_task_class = self.kind_task_class_function(self)      # {类【工件种类，工序】}
        self.time_machine_class_index, self.rate_machine_class_index = self.process_time(self)     # {机器：【类+加工时间】}，{机器：【类+加工速率】}
        self.first_task_class = tuple([self.class_kind[r][0] for r in range(self.number_job_kind)])    # 各类工件首工序类列表
        self.final_task_class = tuple([self.class_kind[r][-1] for r in range(self.number_job_kind)])  # 各类工件末工序类列表

    def process_time(self, data):
        machine_class_dict = {}
        rate_machine_class_dict = {}
        for m in self.machine_list:
            value = []
            rate = []
            for k in self.class_list:
                if m in data.machine_class[k]:
                    value.append(data.time_machine_class[k][data.machine_class[k].index(m)])
                    rate.append(1/data.time_machine_class[k][data.machine_class[k].index(m)])
                else:
                    value.append(None)
                    rate.append(0)
            machine_class_dict[m] = tuple(value)
            rate_machine_class_dict[m] = tuple(rate)
        return machine_class_dict, rate_machine_class_dict

    def class_kind_function(self, data):
        kind_dict = {}
        for i in range(data.number_job_kind):
            class_node_begin = sum(data.number_task[:i])
            class_node_end = sum(data.number_task[:i+1])
            kind_dict[i] = self.class_list[class_node_begin:class_node_end]
        return kind_dict

    def kind_task_class_function(self, data):
        class_dict = {}
        for i in range(data.number_job_kind):
            for j in range(len(self.class_kind[i])):
                class_dict[self.class_kind[i][j]] = tuple([i, j])
        return class_dict




